#!/usr/bin/env python3
import csv, os
_R = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA = os.path.join(_R, "data"); DELIV = os.path.join(_R, "deliverables")
os.makedirs(DELIV, exist_ok=True)
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter

def read(p):
    with open(p) as f:
        return list(csv.reader(f))

thin = Side(style="thin", color="D0D0D0")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
HEADER_FILL = PatternFill("solid", fgColor="1F4E5F")
HEADER_FONT = Font(name="Arial", size=11, bold=True, color="FFFFFF")
BASE_FONT = Font(name="Arial", size=10)

def style_header(ws, ncols, row=1):
    for c in range(1, ncols+1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        cell.border = border
    ws.row_dimensions[row].height = 30

# ============ Schedule ============
sched = read(os.path.join(DATA, "rollout_schedule.csv"))
sh, srows = sched[0], sched[1:]
wb = Workbook()
ws = wb.active
ws.title = "Rollout Schedule"

# Title banner
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(sh))
t = ws.cell(row=1, column=1, value="Worship Music School — Weekly Rollout Schedule (Sept 2026 – Feb 2027)")
t.font = Font(name="Arial", size=14, bold=True, color="1F4E5F")
t.alignment = Alignment(horizontal="left", vertical="center")
ws.row_dimensions[1].height = 26
sub = ws.cell(row=2, column=1, value="Sundays, ~60 min · 3 sections/week: Discipleship (~15) · Musicianship (~20) · Playing Together (~25). Assumptions are adjustable.")
ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(sh))
sub.font = Font(name="Arial", size=9, italic=True, color="666666")

hdr_row = 3
for j, name in enumerate(sh, 1):
    ws.cell(row=hdr_row, column=j, value=name)
style_header(ws, len(sh), row=hdr_row)

FALL = PatternFill("solid", fgColor="FBF3E9")
WINTER = PatternFill("solid", fgColor="EAF2F8")
BREAK = PatternFill("solid", fgColor="EDEDED")
MILE = PatternFill("solid", fgColor="FCE4A6")

for i, r in enumerate(srows):
    xl = hdr_row + 1 + i
    is_break = (r[0] == "—")
    term = r[2]
    for j, val in enumerate(r, 1):
        cell = ws.cell(row=xl, column=j, value=val)
        cell.font = BASE_FONT if not is_break else Font(name="Arial", size=10, italic=True, color="999999")
        cell.border = border
        cell.alignment = Alignment(wrap_text=True, vertical="top",
                                   horizontal="center" if j in (1,2,3) else "left")
        if is_break:
            cell.fill = BREAK
        elif term == "Fall":
            cell.fill = FALL
        elif term == "Winter":
            cell.fill = WINTER
    # milestone highlight on notes cell
    notes = r[7]
    if any(k in notes for k in ("MILESTONE","SHOWCASE","LAUNCH")):
        nc = ws.cell(row=xl, column=8)
        nc.fill = MILE
        nc.font = Font(name="Arial", size=10, bold=True)

widths = [6, 12, 8, 30, 26, 34, 34, 30]
for j, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(j)].width = w
ws.freeze_panes = f"A{hdr_row+1}"
ws.auto_filter.ref = f"A{hdr_row}:{get_column_letter(len(sh))}{hdr_row+len(srows)}"
ws.sheet_view.showGridLines = False
wb.save(os.path.join(DELIV, "Worship Music School - Rollout Schedule.xlsx"))
print("Saved schedule xlsx")

# ============ Song Library ============
lib = read(os.path.join(DATA, "song_library.csv"))
lh, lrows = lib[0], lib[1:]
wb2 = Workbook()
ws2 = wb2.active
ws2.title = "Song Library"

ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(lh))
t2 = ws2.cell(row=1, column=1, value="Worship Music School — Song Library & Difficulty Guide")
t2.font = Font(name="Arial", size=14, bold=True, color="1F4E5F")
ws2.row_dimensions[1].height = 26
ws2.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(lh))
leg = ws2.cell(row=2, column=1, value="Difficulty per instrument: 1 = very easy · 2 = easy · 3 = intermediate · 4 = advanced · 5 = hard.  Green = easier, red = harder. Keys are band-friendly starting points — transpose to your vocalist.")
leg.font = Font(name="Arial", size=9, italic=True, color="666666")
leg.alignment = Alignment(wrap_text=True, vertical="center")
ws2.row_dimensions[2].height = 26

hdr2 = 3
for j, name in enumerate(lh, 1):
    ws2.cell(row=hdr2, column=j, value=name)
style_header(ws2, len(lh), row=hdr2)

idx = {c:i for i,c in enumerate(lh)}
diff_cols = ["Drums","Bass","Guitar","Keys","Vocals","Overall"]
type_fill = {"Worship": "F2F7F4", "Christmas": "F7EFF5", "Secular": "F3F4F7"}

for i, r in enumerate(lrows):
    xl = hdr2 + 1 + i
    tval = r[idx["Type"]]
    for j, val in enumerate(r, 1):
        # convert numeric difficulty to int
        cell = ws2.cell(row=xl, column=j, value=int(val) if lh[j-1] in diff_cols and val.isdigit() else val)
        cell.font = BASE_FONT
        cell.border = border
        center = (lh[j-1] in diff_cols) or lh[j-1] in ("BPM","Key")
        cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="center" if center else "left")
        cell.fill = PatternFill("solid", fgColor=type_fill.get(tval, "FFFFFF"))

# color scale on difficulty columns
first_data = hdr2 + 1
last_data = hdr2 + len(lrows)
for name in diff_cols:
    col = get_column_letter(idx[name]+1)
    rng = f"{col}{first_data}:{col}{last_data}"
    rule = ColorScaleRule(start_type="num", start_value=1, start_color="63BE7B",
                          mid_type="num", mid_value=3, mid_color="FFEB84",
                          end_type="num", end_value=5, end_color="F8696B")
    ws2.conditional_formatting.add(rng, rule)

widths2 = {"Title":30,"Artist":22,"Type":11,"Key":6,"BPM":6,"Feel":18,
           "Drums":7,"Bass":7,"Guitar":8,"Keys":7,"Vocals":8,"Overall":8,
           "Best For":24,"Teaching Focus":30,"Notes":34}
for name, w in widths2.items():
    ws2.column_dimensions[get_column_letter(idx[name]+1)].width = w
ws2.freeze_panes = f"D{hdr2+1}"   # keep Title/Artist/Type visible while scrolling
ws2.auto_filter.ref = f"A{hdr2}:{get_column_letter(len(lh))}{last_data}"
ws2.sheet_view.showGridLines = False
wb2.save(os.path.join(DELIV, "Worship Music School - Song Library.xlsx"))
print("Saved song library xlsx")
